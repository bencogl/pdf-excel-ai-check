from fastapi import FastAPI, UploadFile, File, HTTPException
from huggingface_hub import InferenceClient
import pdfplumber
from openpyxl import load_workbook
import traceback

app = FastAPI()

HUGGINGFACE_TOKEN = "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"
model_name = "mistralai/Mistral-7B-Instruct-v0.1"

client = InferenceClient(model=model_name, token=HUGGINGFACE_TOKEN)

@app.get("/")
def home():
    return {"messaggio": "Carica PDF ed Excel tramite endpoint /upload/"}

@app.post("/upload/")
async def analizza_documenti(pdf: UploadFile = File(...), excel: UploadFile = File(...)):
    try:
        # Estrazione testo dal PDF
        testo_pdf = ""
        with pdfplumber.open(pdf.file) as pdf_file:
            for pagina in pdf_file.pages:
                estratto = pagina.extract_text()
                if estratto:
                    testo_pdf += estratto
                else:
                    testo_pdf += "[Pagina vuota o non leggibile]\n"

        # Estrazione dati dall'Excel
        wb = load_workbook(excel.file, data_only=True)
        foglio = wb.active
        dati_excel = []
        for riga in foglio.iter_rows(values_only=True):
            dati_excel.append(riga)

        # Creazione del prompt per l'AI
        prompt_ai = f"""
        Dati estratti dal PDF:
        {testo_pdf}

        Dati estratti dall'Excel:
        {dati_excel}

        Identifica chiaramente eventuali differenze, incongruenze o errori.
        """

        # Chiamata all'API di HuggingFace
        risposta_ai = client.text_generation(prompt_ai, max_new_tokens=200)
        return {"Analisi effettuata da AI": risposta_ai}

    except Exception as e:
        # Log dettagliato dell'errore
        errore_dettagliato = traceback.format_exc()
        print(f"Errore interno: {errore_dettagliato}")
        raise HTTPException(status_code=500, detail=errore_dettagliato)
